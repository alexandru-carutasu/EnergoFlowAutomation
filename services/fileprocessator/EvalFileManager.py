"""
EvalFileManager - Service for creating and updating evaluation files.

Handles evaluation file operations including creation from templates,
data merging, and uploading to Dropbox.
"""

import logging
import os
import tempfile
from datetime import datetime as dt, timedelta
from typing import Optional, List, Tuple

import pandas as pd
import pytz
from openpyxl import load_workbook

from config import FORECAST_TAG, IBD_TAG
from services.dropboxclient.DropboxClient import DropboxClient
from services.dbmanager.DbManager import DbManager

logger = logging.getLogger(__name__)


class EvalFileManager:
    """Manages evaluation file creation and updates."""

    LOCAL_TZ = pytz.timezone("Europe/Bucharest")
    UTC_TZ = pytz.utc

    EVAL_FILE_PREFIX = "Evaluare_"
    EVAL_FILE_SUFFIX = ".xlsx"
    EVAL_FOLDER_TEMPLATE = "/Forecast Automat/{client_name}"
    EVAL_FOLDER_OUT = "/Forecast Automat/{client_name}/out"  # For forecast
    EVAL_FOLDER_IN = "/Forecast Automat/{client_name}/in"    # For production (IBD)
    TEMPLATE_FOLDER = "/Forecast Automat"
    TEMPLATE_FILE_NAME = "Evaluare.xlsx"

    def __init__(self, db_manager: DbManager, dropbox_client: DropboxClient) -> None:
        self.db_manager = db_manager
        self.dropbox_client = dropbox_client

    def _convert_utc_to_local(self, utc_date, utc_interval: str) -> Tuple[any, str]:
        """Convert UTC date and interval to local time (Europe/Bucharest).

        Args:
            utc_date: Date object (UTC)
            utc_interval: Time interval string like "08:00" (UTC)

        Returns:
            Tuple of (local_date, local_interval_string)
        """
        if isinstance(utc_date, dt):
            base_date = utc_date.date()
        else:
            base_date = utc_date

        interval_time = dt.strptime(utc_interval, "%H:%M").time()
        utc_naive = dt.combine(base_date, interval_time)
        utc_dt = self.UTC_TZ.localize(utc_naive)
        local_dt = utc_dt.astimezone(self.LOCAL_TZ)

        return local_dt.date(), local_dt.strftime("%H:%M")

    def create_evaluation_file(
        self, client_id: int, client_name: str, date_str: str
    ) -> bool:
        logging.info(f"Creating evaluation file for {client_name} - {date_str}")
        try:
            # Get plant names from database
            plants = self.db_manager.get_plants_by_client(client_id)
            if not plants:
                logging.warning(f"No plants found for client {client_id}")
                return False

            plant_names = [plant.name for plant in plants]

            # Download template from Dropbox
            template_path = self._download_template()
            if not template_path:
                logging.error(f"Failed to download template from Dropbox {template_path}.")
                return False

            # Create evaluation file from template
            local_file_name = f"{self.EVAL_FILE_PREFIX}{client_name}_{date_str}{self.EVAL_FILE_SUFFIX}"
            eval_file_path = self._create_eval_file_from_template(
                template_path, local_file_name, plant_names, date_str
            )

            if not eval_file_path:
                logging.error("Failed to create evaluation file")
                if os.path.exists(template_path):
                    os.remove(template_path)
                return False

            # Upload to Dropbox
            dropbox_folder = self.EVAL_FOLDER_TEMPLATE.format(client_name=client_name)
            upload_result = self.dropbox_client.upload_excel_to_dropbox(
                eval_file_path, dropbox_folder
            )

            # Cleanup
            try:
                os.remove(template_path)
                os.remove(eval_file_path)
            except OSError as e:
                logging.warning(f"Error cleaning up files: {e}")

            if not upload_result:
                logging.error(f"Failed to upload evaluation file to Dropbox")
                return False

            logging.info(f"Successfully created evaluation file for {client_name}")
            return True

        except Exception as e:
            logging.error(f"Error creating evaluation file: {e}", exc_info=True)
            return False

    def update_evaluation_files(
        self, client_name: str, client_id: int, curr_date: dt, tag: str
    ) -> bool:
        """Update evaluation files for multi-plant client.

        Parameters
        ----------
        client_name : str
            Client name
        client_id : int
            Client ID
        curr_date : datetime
            Current date
        tag : str
            Tag indicating file type (FORECAST_TAG or IBD_TAG)

        Returns
        -------
        bool
            True if successful
        """
        logging.info(f"Updating evaluation files for {client_name} (tag: {tag})")
        try:
            # Download existing evaluation file or create new one
            formatted_date = curr_date.strftime("%m%Y")
            eval_file_path = self._download_eval_file(client_name, formatted_date)

            if not eval_file_path or not os.path.exists(eval_file_path):
                logging.info(f"Evaluation file not found, creating new one for {client_name}")
                if not self.create_evaluation_file(client_id, client_name, formatted_date):
                    logging.error(f"Failed to create evaluation file for {client_name}")
                    return False
                eval_file_path = self._download_eval_file(client_name, formatted_date)
                if not eval_file_path or not os.path.exists(eval_file_path):
                    logging.error(f"Still failed to get evaluation file after creation")
                    return False

            # Get all plants for client
            plants = self.db_manager.get_plants_by_client(client_id)

            for plant in plants:
                self._update_plant_sheet(eval_file_path, plant.id, plant.name, curr_date)

            # Evaluation file always goes to client root folder
            dropbox_folder = self.EVAL_FOLDER_TEMPLATE.format(client_name=client_name)

            target_filename = f"{self.EVAL_FILE_PREFIX}{client_name}_{formatted_date}{self.EVAL_FILE_SUFFIX}"
            upload_result = self.dropbox_client.upload_excel_to_dropbox(
                eval_file_path, dropbox_folder, target_filename
            )

            if upload_result:
                logging.info(f"Successfully updated evaluation file for {client_name} to {dropbox_folder}")
                try:
                    os.remove(eval_file_path)
                except OSError:
                    pass
                return True
            else:
                logging.error("Failed to upload updated evaluation file to Dropbox")
                return False

        except Exception as e:
            logging.error(f"Error updating evaluation files: {e}", exc_info=True)
            return False

    def update_evaluation_files_1plant(
        self, client_name: str, client_id: int, plant_name: str, curr_date: dt
    ) -> bool:
        logging.info(
            f"Updating evaluation file for {client_name} - plant {plant_name}"
        )
        try:
            # Download existing evaluation file
            formatted_date = curr_date.strftime("%m%Y")
            eval_file_path = self._download_eval_file(client_name, formatted_date)

            if not eval_file_path or not os.path.exists(eval_file_path):
                logging.error(f"Failed to download evaluation file from Dropbox")
                return False

            # Get plant ID
            plants = self.db_manager.get_plants_by_client(client_id)
            plant_id = None
            for plant in plants:
                if plant.name == plant_name:
                    plant_id = plant.id
                    break

            if plant_id is None:
                logging.error(f"Plant {plant_name} not found for client {client_id}")
                return False

            # Update sheet for this plant
            self._update_plant_sheet(eval_file_path, plant_id, plant_name, curr_date)

            # Upload updated file
            dropbox_folder = self.EVAL_FOLDER_TEMPLATE.format(client_name=client_name)
            upload_result = self.dropbox_client.upload_excel_to_dropbox(
                eval_file_path, dropbox_folder
            )

            if upload_result:
                logging.info(
                    f"Successfully updated evaluation file for {plant_name}"
                )
                try:
                    os.remove(eval_file_path)
                except OSError:
                    pass
                return True
            else:
                logging.error("Failed to upload updated evaluation file to Dropbox")
                return False

        except Exception as e:
            logging.error(f"Error updating evaluation file for 1 plant: {e}", exc_info=True)
            return False

    def update_evaluation_file(
        self, client_name: str, client_id: int, curr_date: dt
    ) -> bool:
        """Update evaluation file by updating all plants.

        Parameters
        ----------
        client_name : str
            Client name
        client_id : int
            Client ID
        curr_date : datetime
            Current date

        Returns
        -------
        bool
            True if successful
        """
        logging.info(f"Updating evaluation file for {client_name}")
        try:
            plants = self.db_manager.get_plants_by_client(client_id)
            if not plants:
                logging.warning(f"No plants found for client {client_id}")
                return False

            for plant in plants:
                self.update_evaluation_files_1plant(
                    client_name, client_id, plant.name, curr_date
                )

            return True
        except Exception as e:
            logging.error(f"Error updating evaluation file: {e}", exc_info=True)
            return False

    # ─────────────────────────────────────────────────────────────────
    # Helper Methods
    # ─────────────────────────────────────────────────────────────────

    def _download_template(self) -> Optional[str]:
        """Download template evaluation file from Dropbox."""
        try:
            refresh_token = self.dropbox_client.load_refresh_token()
            if not refresh_token:
                logging.error("Refresh token not found")
                return None

            dbx = self.dropbox_client.get_access_token(refresh_token)
            if not dbx:
                logging.error("Failed to authenticate with Dropbox")
                return None

            # Create temporary file
            with tempfile.NamedTemporaryFile(
                suffix=".xlsx", delete=False
            ) as tmp_file:
                tmp_path = tmp_file.name

            try:
                template_dropbox_path = f"{self.TEMPLATE_FOLDER}/{self.TEMPLATE_FILE_NAME}"
                metadata, response = dbx.files_download(template_dropbox_path)

                with open(tmp_path, "wb") as f:
                    f.write(response.content)

                logging.info(f"Downloaded template from {template_dropbox_path}")
                return tmp_path
            except Exception as e:
                logging.error(f"Error downloading template from Dropbox: {e}")
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)
                return None

        except Exception as e:
            logging.error(f"Error in _download_template: {e}")
            return None

    def _download_eval_file(self, client_name: str, date_str: str) -> Optional[str]:
        """Download existing evaluation file from Dropbox."""
        try:
            refresh_token = self.dropbox_client.load_refresh_token()
            if not refresh_token:
                logging.error("Refresh token not found")
                return None

            dbx = self.dropbox_client.get_access_token(refresh_token)
            if not dbx:
                logging.error("Failed to authenticate with Dropbox")
                return None

            # Create temporary file
            with tempfile.NamedTemporaryFile(
                suffix=".xlsx", delete=False
            ) as tmp_file:
                tmp_path = tmp_file.name

            try:
                dropbox_path = f"{self.EVAL_FOLDER_TEMPLATE.format(client_name=client_name)}/{self.EVAL_FILE_PREFIX}{client_name}_{date_str}{self.EVAL_FILE_SUFFIX}"
                metadata, response = dbx.files_download(dropbox_path)

                with open(tmp_path, "wb") as f:
                    f.write(response.content)

                logging.info(f"Downloaded evaluation file from {dropbox_path}")
                return tmp_path
            except Exception as e:
                logging.error(f"Error downloading evaluation file from Dropbox: {e}")
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)
                return None

        except Exception as e:
            logging.error(f"Error in _download_eval_file: {e}")
            return None

    def _create_eval_file_from_template(
        self, template_path: str, output_file_name: str, plant_names: List[str], date_str: str
    ) -> Optional[str]:
        """Create evaluation file from template.

        Parameters
        ----------
        template_path : str
            Path to template file
        output_file_name : str
            Output file name
        plant_names : List[str]
            List of plant names (sheet names)
        date_str : str
            Date in MMYYYY format (e.g., "052026")

        Returns
        -------
        Optional[str]
            Path to created file or None
        """
        try:
            workbook = load_workbook(template_path)
            logging.info(f"Template sheets: {workbook.sheetnames}")

            if len(workbook.sheetnames) < 2:
                logging.error("Template must have at least 2 sheets (summary + 1 plant template)")
                return None

            num_plants = len(plant_names)
            # Sheet 0 = Summary (keep), Sheets 1 to num_plants = plant sheets (rename), rest = delete
            sheets_to_keep = 1 + num_plants  # summary + plant sheets

            # Delete excess sheets from the end
            while len(workbook.sheetnames) > sheets_to_keep:
                sheet_to_remove = workbook.sheetnames[-1]
                workbook.remove(workbook[sheet_to_remove])
                logging.info(f"Removed excess sheet: {sheet_to_remove}")

            # Rename plant sheets (sheets 1 to num_plants) to plant names from DB
            for i, plant_name in enumerate(plant_names):
                sheet_index = i + 1  # skip summary at index 0
                if sheet_index < len(workbook.sheetnames):
                    old_name = workbook.sheetnames[sheet_index]
                    workbook[old_name].title = plant_name
                    logging.info(f"Renamed sheet '{old_name}' to '{plant_name}'")

            # Populate date and interval columns for each plant sheet
            self._populate_date_interval_columns(workbook, plant_names, date_str)

            # Save file
            workbook.save(output_file_name)
            logging.info(f"Created evaluation file: {output_file_name} with sheets: {workbook.sheetnames}")
            return output_file_name

        except Exception as e:
            logging.error(f"Error creating evaluation file from template: {e}")
            return None

    def _populate_date_interval_columns(
        self, workbook, plant_names: List[str], date_str: str
    ) -> None:
        """Populate date and interval columns for each plant sheet.

        Creates rows for every 15-minute interval of the month.

        Parameters
        ----------
        workbook : Workbook
            openpyxl workbook object
        plant_names : List[str]
            List of plant sheet names
        date_str : str
            Date in MMYYYY format (e.g., "052026")
        """
        import calendar

        # Parse month and year from date_str (MMYYYY format)
        month = int(date_str[:2])
        year = int(date_str[2:])

        # Get number of days in month
        _, days_in_month = calendar.monthrange(year, month)

        # Generate all 15-minute intervals for the month
        intervals = []
        for day in range(1, days_in_month + 1):
            date_val = dt(year, month, day).date()
            for hour in range(24):
                for minute in [0, 15, 30, 45]:
                    time_str = f"{hour:02d}:{minute:02d}"
                    intervals.append((date_val, time_str))

        logging.info(f"Generating {len(intervals)} rows for {month}/{year} ({days_in_month} days)")

        # Populate each plant sheet
        for plant_name in plant_names:
            if plant_name not in workbook.sheetnames:
                logging.warning(f"Sheet {plant_name} not found in workbook")
                continue

            sheet = workbook[plant_name]
            start_row = 4  # Data starts at row 4

            for idx, (date_val, time_str) in enumerate(intervals):
                row = start_row + idx
                sheet.cell(row=row, column=1, value=date_val)
                sheet.cell(row=row, column=2, value=time_str)

            logging.info(f"Populated {len(intervals)} date/interval rows in sheet {plant_name}")

    def _ensure_date_interval_populated(
        self, file_path: str, sheet_name: str, curr_date: dt
    ) -> None:
        """Ensure date and interval columns are populated in a sheet.

        If row 4 column 1 is empty, populates the entire month's dates/intervals.

        Parameters
        ----------
        file_path : str
            Path to Excel file
        sheet_name : str
            Sheet name to check/populate
        curr_date : datetime
            Current date (used to determine month)
        """
        import calendar

        try:
            workbook = load_workbook(file_path)

            if sheet_name not in workbook.sheetnames:
                logging.warning(f"Sheet {sheet_name} not found for date/interval population")
                return

            sheet = workbook[sheet_name]

            # Check if dates are already populated (row 4, column 1)
            if sheet.cell(row=4, column=1).value is not None:
                logging.info(f"Sheet {sheet_name} already has date/interval data")
                return

            # Get month/year from current date
            month = curr_date.month
            year = curr_date.year
            _, days_in_month = calendar.monthrange(year, month)

            # Generate all 15-minute intervals for the month
            intervals = []
            for day in range(1, days_in_month + 1):
                date_val = dt(year, month, day).date()
                for hour in range(24):
                    for minute in [0, 15, 30, 45]:
                        time_str = f"{hour:02d}:{minute:02d}"
                        intervals.append((date_val, time_str))

            logging.info(f"Populating {len(intervals)} date/interval rows in sheet {sheet_name}")

            start_row = 4
            for idx, (date_val, time_str) in enumerate(intervals):
                row = start_row + idx
                sheet.cell(row=row, column=1, value=date_val)
                sheet.cell(row=row, column=2, value=time_str)

            workbook.save(file_path)
            logging.info(f"Saved sheet {sheet_name} with populated date/interval columns")

        except Exception as e:
            logging.error(f"Error ensuring date/interval populated for {sheet_name}: {e}")

    def _update_plant_sheet(
        self, eval_file_path: str, plant_id: int, plant_name: str, curr_date: dt
    ) -> bool:
        """Update sheet for specific plant with measurements and prices.

        Parameters
        ----------
        eval_file_path : str
            Path to evaluation file
        plant_id : int
            Plant ID
        plant_name : str
            Plant name (sheet name)
        curr_date : datetime
            Current date

        Returns
        -------
        bool
            True if successful
        """
        try:
            logging.info(f"Updating sheet for plant: {plant_name}")

            # Check if date/interval columns need to be populated
            self._ensure_date_interval_populated(eval_file_path, plant_name, curr_date)

            # Get measurements for plant from start of month
            start_of_month = curr_date.replace(day=1)
            start_date = start_of_month.date()

            measurements = self.db_manager.get_measurements_by_plant_from_date(
                plant_id, start_date
            )

            logging.info(f"Found {len(measurements)} measurements for plant {plant_name}")

            if not measurements:
                logging.warning(
                    f"No measurements found for plant {plant_name} from {start_date}"
                )
                return False

            # Create dataframe from measurements
            meas_data = []
            for m in measurements:
                # Parse date string to date object if needed
                m_date = m.date
                if isinstance(m_date, str):
                    m_date = dt.strptime(m_date, "%Y-%m-%d").date()

                # Convert UTC to local time for matching with Excel file
                local_date, local_interval = self._convert_utc_to_local(m_date, m.interval)

                meas_data.append(
                    {
                        "data": local_date,
                        "timp": local_interval,
                        "productionmw": m.prod_val,
                        "forecastmw": m.forecast_val,
                    }
                )

            dfMeas = pd.DataFrame(meas_data)
            logging.info(f"Measurements dataframe shape: {dfMeas.shape}, dates: {dfMeas['data'].unique()[:5]}")

            # Get imbalance prices for current month
            df_prices = self._get_imbalance_prices(start_date)

            # Merge measurements with prices
            dfAll = pd.merge(dfMeas, df_prices, on=["data", "timp"], how="outer")
            dfAll = dfAll.sort_values(by=["data", "timp"])

            # Keep only current month data
            current_month_str = curr_date.strftime("%Y-%m")
            dfAll = dfAll[dfAll["data"].astype(str).str.startswith(current_month_str)]

            # Write data to sheet
            self._write_data_to_sheet(eval_file_path, plant_name, dfAll)

            return True

        except Exception as e:
            logging.error(f"Error updating plant sheet {plant_name}: {e}", exc_info=True)
            return False

    def _get_imbalance_prices(self, start_date) -> pd.DataFrame:
        """Get imbalance prices from start date onwards.

        Parameters
        ----------
        start_date : date
            Start date filter

        Returns
        -------
        pd.DataFrame
            DataFrame with prices
        """
        try:
            prices = self.db_manager.get_all_imbalance_prices()

            if not prices:
                return pd.DataFrame(columns=["data", "timp", "positive_price", "negative_price"])

            price_data = []
            for price in prices:
                # Parse date string to date object if needed
                price_date = price.date
                if isinstance(price_date, str):
                    price_date = dt.strptime(price_date, "%Y-%m-%d").date()

                # Convert UTC to local time for matching with Excel file
                local_date, local_interval = self._convert_utc_to_local(price_date, price.interval)

                price_data.append(
                    {
                        "data": local_date,
                        "timp": local_interval,
                        "positive_price": price.positive_imbalance,
                        "negative_price": price.negative_imbalance,
                    }
                )

            df_prices = pd.DataFrame(price_data)
            df_prices = df_prices[df_prices["data"] >= start_date]

            return df_prices
        except Exception as e:
            logging.error(f"Error getting imbalance prices: {e}")
            return pd.DataFrame(
                columns=["data", "timp", "positive_price", "negative_price"]
            )

    def _write_data_to_sheet(
        self, file_name: str, sheet_name: str, df_all: pd.DataFrame
    ) -> bool:
        """Write measurement and price data to Excel sheet.

        Parameters
        ----------
        file_name : str
            Excel file path
        sheet_name : str
            Sheet name
        df_all : pd.DataFrame
            DataFrame with data to write

        Returns
        -------
        bool
            True if successful
        """
        try:
            workbook = load_workbook(file_name)

            if sheet_name not in workbook.sheetnames:
                logging.warning(f"Sheet {sheet_name} not found in workbook")
                return False

            sheet = workbook[sheet_name]
            start_row = 4  # Data starts at row 4

            # Log first few rows to debug
            logging.info(f"Sheet {sheet_name} row 4 col 1: {sheet.cell(row=4, column=1).value}, col 2: {sheet.cell(row=4, column=2).value}")
            logging.info(f"DataFrame columns: {df_all.columns.tolist()}, shape: {df_all.shape}")
            if not df_all.empty:
                logging.info(f"DataFrame first row: {df_all.iloc[0].to_dict()}")

            rows_updated = 0
            while (
                sheet.cell(row=start_row, column=1).value is not None
                and str(sheet.cell(row=start_row, column=1).value).strip() != "TOTAL LUNA"
                and start_row < 3000
            ):
                # Get date and time from sheet
                sheet_date = sheet.cell(row=start_row, column=1).value
                timp_cell = sheet.cell(row=start_row, column=2).value

                if not sheet_date or not timp_cell:
                    start_row += 1
                    continue

                timp = str(timp_cell)[:5]  # HH:MM format

                # Normalize sheet date to date object for comparison
                if isinstance(sheet_date, dt):
                    sheet_date_normalized = sheet_date.date()
                elif hasattr(sheet_date, 'date'):
                    sheet_date_normalized = sheet_date
                else:
                    sheet_date_normalized = dt.strptime(str(sheet_date)[:10], "%Y-%m-%d").date()

                # Find matching row in dataframe using normalized date comparison
                matching_rows = df_all[
                    (df_all["data"] == sheet_date_normalized) & (df_all["timp"] == timp)
                ]

                if not matching_rows.empty:
                    row_data = matching_rows.iloc[0]
                    rows_updated += 1

                    # Write production
                    prod_val = row_data.get("productionmw")
                    if prod_val is not None and prod_val != sheet.cell(
                        row=start_row, column=3
                    ).value:
                        sheet.cell(row=start_row, column=3, value=prod_val)

                    # Write forecast
                    fcst_val = row_data.get("forecastmw")
                    if fcst_val is not None and fcst_val != sheet.cell(
                        row=start_row, column=6
                    ).value:
                        sheet.cell(row=start_row, column=6, value=fcst_val)

                    # Write positive price
                    pos_price = row_data.get("positive_price")
                    if (
                        pos_price is not None
                        and pos_price != sheet.cell(row=start_row, column=9).value
                    ):
                        sheet.cell(row=start_row, column=9, value=pos_price)

                    # Write negative price
                    neg_price = row_data.get("negative_price")
                    if (
                        neg_price is not None
                        and neg_price != sheet.cell(row=start_row, column=10).value
                    ):
                        sheet.cell(row=start_row, column=10, value=neg_price)

                start_row += 1

            logging.info(f"Updated {rows_updated} rows in sheet {sheet_name}")

            workbook.save(file_name)
            logging.info(f"Successfully updated sheet {sheet_name}")
            return True

        except Exception as e:
            logging.error(f"Error writing data to sheet: {e}", exc_info=True)
            return False
