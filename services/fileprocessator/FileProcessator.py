import logging
import os
import tempfile
import datetime
import pytz
import imaplib
from typing import Iterable, Tuple, Any, Optional
from openpyxl import load_workbook
from datetime import datetime as dt, timedelta

from config import (
    DROPBOX_APP_KEY,
    DROPBOX_APP_SECRET,
    DROPBOX_TOKEN_FILE,
    DROPBOX_EVAL_FILE_PATH,
    PROCESSING_TAG,
    DONE_TAG,
    IBD_TAG,
    FORECAST_TAG,
    IMAP_SERVER,
    IMAP_USER,
    IMAP_PASSWORD,
)
from services.dropboxclient.DropboxClient import DropboxClient
from services.dbmanager.DbManager import DbManager
from services.fileprocessator.EvalFileManager import EvalFileManager

logger = logging.getLogger(__name__)


class FileProcessator:
    """
    Advanced service for processing XLSX files with database integration.

    Each item in xlsx_files is expected to be a tuple:
        (file_name, payload, curr_email, tag, sender, email_timestamp, client_name)
    """

    def __init__(self, db_manager: DbManager) -> None:
        """Initialize FileProcessator with database and Dropbox clients."""
        self.xlsx_files: list[Tuple[str, bytes, Any, str, str, Any, str]] = []
        self.db_manager = db_manager
        self.dropbox_client = DropboxClient(
            DROPBOX_APP_KEY,
            DROPBOX_APP_SECRET,
            DROPBOX_TOKEN_FILE,
            DROPBOX_EVAL_FILE_PATH,
        )
        self.eval_file_manager = EvalFileManager(db_manager, self.dropbox_client)

    def set_xlsx_files(
        self, xlsx_files: Iterable[Tuple[str, bytes, Any, str, str, Any, str]]
    ) -> None:
        """Set XLSX files for processing."""
        self.xlsx_files = list(xlsx_files)

    def process_xlsx_files(self) -> None:
        """Process all XLSX files through the complete workflow."""
        logging.info("Processing %d XLSX file(s)...", len(self.xlsx_files))

        if not self.xlsx_files:
            return

        for file_name, payload, curr_email, tag, sender, email_timestamp, client_name in self.xlsx_files:
            try:
                self._process_single_file(
                    file_name, payload, curr_email, tag, sender, email_timestamp, client_name
                )
            except Exception as e:
                logging.error(f"Error processing file {file_name}: {e}", exc_info=True)

    def _process_single_file(
        self,
        file_name: str,
        payload: bytes,
        curr_email: Any,
        tag: str,
        sender: str,
        email_timestamp: Any,
        client_name: str,
    ) -> None:
        """Process a single XLSX file through the complete workflow."""
        client_id = None
        if not client_name and sender:
            client = self.db_manager.get_client_by_email(sender)
            if client:
                client_name = client.name
                client_id = client.id
            else:
                logging.warning(f"No client found for sender {sender}. Skipping file {file_name}.")
                return

        local_file_path = file_name
        with open(local_file_path, "wb") as f:
            f.write(payload)

        try:
            # Store measurements in database with tag
            result = self._store_measurements_into_db(local_file_path, client_id, tag)
            if result:
                logging.error(f"Failed to store measurements: {result}")
            else:
                # Upload original email file to /out (forecast) or /in (production)
                self._upload_email_file_to_dropbox(local_file_path, file_name, client_name, tag)

                # Update evaluation files on Dropbox after successful DB storage
                curr_date = dt.now()
                self._update_evaluation_files(client_name, client_id, curr_date, tag)
        except Exception as e:
            logging.error(f"Error processing file {file_name}: {e}", exc_info=True)
        finally:
            try:
                os.remove(local_file_path)
            except OSError:
                pass

    def _upload_email_file_to_dropbox(
        self, local_file_path: str, original_filename: str, client_name: str, tag: str
    ) -> None:
        """Upload the original email attachment to Dropbox /out or /in folder."""
        try:
            if tag == FORECAST_TAG:
                dropbox_folder = f"/Forecast Automat/{client_name}/out"
            elif tag == IBD_TAG:
                dropbox_folder = f"/Forecast Automat/{client_name}/in"
            else:
                logging.warning(f"Unknown tag {tag}, skipping Dropbox upload")
                return

            result = self.dropbox_client.upload_excel_to_dropbox(
                local_file_path, dropbox_folder, original_filename
            )
            if result:
                logging.info(f"Uploaded {original_filename} to {dropbox_folder}")
            else:
                logging.error(f"Failed to upload {original_filename} to Dropbox")
        except Exception as e:
            logging.error(f"Error uploading email file to Dropbox: {e}")

    def _store_measurements_into_db(
        self, file_path: str, client_id: Optional[int], tag: str
    ) -> str:
        """Store measurements from XLSX into database with tag support."""
        logging.info("Storing measurements...")
        try:
            workbook, header, forecast_idx = self._load_workbook_and_header(file_path)
            if workbook is None:
                return "ERROR"

            plants = self.db_manager.get_plants_by_client(client_id) if client_id else []
            plant_map = {p.name: p.id for p in plants}

            for sheet_name in workbook.sheetnames:
                plant_id = plant_map.get(sheet_name)
                if plant_id is None:
                    logging.warning(f"No plant found for sheet '{sheet_name}' (client_id={client_id}). Skipping sheet.")
                    continue
                error = self._process_sheet_rows(
                    workbook, sheet_name, plant_id, header, forecast_idx, tag
                )
                if error:
                    return error

            return ""
        except Exception as e:
            logging.error(f"Error storing measurements: {e}")
            return "ERROR"

    def _load_workbook_and_header(self, file_path: str) -> Tuple[Any, dict, Optional[int]]:
        """Load workbook and extract header with forecast column index."""
        try:
            workbook = load_workbook(file_path)
            sheet_names = workbook.sheetnames
            if not sheet_names:
                raise ValueError("No sheets found in workbook")

            # Access first sheet for header
            sheet = workbook[sheet_names[0]]
            header = {cell.value: idx for idx, cell in enumerate(sheet[1])}

            # Find forecast column
            forecast_idx = self._find_forecast_column(header)
            if forecast_idx is None:
                raise ValueError("Forecast column not found in sheet")

            return workbook, header, forecast_idx
        except Exception as e:
            logging.error(f"Error loading workbook: {e}")
            return None, {}, None

    def _find_forecast_column(self, header: dict) -> Optional[int]:
        """Find forecast column index from header row."""
        forecast_column_names = ["P FINAL", "Prognoza la 15 minute", "PROGNOZA"]
        for col_name in forecast_column_names:
            if col_name in header:
                return header[col_name]
        return None

    def _process_sheet_rows(
        self,
        workbook: Any,
        sheet_name: str,
        plant_id: int,
        header: dict,
        forecast_idx: int,
        tag: str,
    ) -> str:
        """Process all rows in a sheet and store measurements."""
        try:
            logging.info(f"Processing sheet: {sheet_name}")

            for row in workbook[sheet_name].iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue

                measurements, _ = self._parse_row_and_add_measurements(
                    row, header, forecast_idx, tag
                )
                for m in measurements:
                    self.db_manager.upsert_measurement(
                        plant_id=plant_id,
                        date_=m["date"],
                        interval=m["interval"],
                        forecast_val=m.get("forecast_val"),
                        prod_val=m.get("prod_val"),
                    )

            return ""
        except Exception as e:
            logging.error(f"Error processing sheet {sheet_name}: {e}")
            return "SHEET_ERROR"

    def _parse_row_and_add_measurements(
        self, row: Tuple, header: dict, forecast_idx: int, tag: str
    ) -> Tuple[list, Optional[Any]]:
        """Parse a single row and create measurement entries with tag-based column selection.

        Converts local time (Europe/Bucharest) to UTC before storing.
        """
        try:
            date = row[header.get("ZIUA", 0)]
            hour_interval = row[header.get("INTERVAL", 1)]
            data_value = row[forecast_idx]

            if not date or not hour_interval:
                return [], None

            intervals = self._parse_time_intervals(hour_interval)
            measurements = []

            # Timezone for conversion
            local_tz = pytz.timezone("Europe/Bucharest")

            for interval_start in intervals:
                # Convert local datetime to UTC
                date_utc, interval_utc = self._convert_to_utc(date, interval_start, local_tz)

                # Unified storage: use tag to determine which column to populate
                if tag == FORECAST_TAG:
                    measurement = {
                        "date": date_utc,
                        "interval": interval_utc,
                        "forecast_val": data_value,
                    }
                elif tag == IBD_TAG:
                    measurement = {
                        "date": date_utc,
                        "interval": interval_utc,
                        "prod_val": data_value,
                    }
                else:
                    logging.warning(f"Unknown tag: {tag}. Skipping.")
                    continue

                measurements.append(measurement)

            return measurements, date
        except Exception as e:
            logging.error(f"Error parsing row: {e}")
            return [], None

    def _convert_to_utc(self, date_val: Any, interval: str, local_tz) -> Tuple[Any, str]:
        """Convert local date and interval to UTC.

        Args:
            date_val: Date from Excel (can be datetime.date or datetime.datetime)
            interval: Time interval string like "08:00"
            local_tz: pytz timezone object

        Returns:
            Tuple of (utc_date, utc_interval_string)
        """
        # Handle both date and datetime from Excel
        if isinstance(date_val, dt):
            base_date = date_val.date()
        elif isinstance(date_val, datetime.date):
            base_date = date_val
        else:
            # Try parsing as string
            base_date = dt.strptime(str(date_val), "%Y-%m-%d").date()

        # Parse interval time
        interval_time = dt.strptime(interval, "%H:%M").time()

        # Combine date and time into naive datetime
        local_naive = dt.combine(base_date, interval_time)

        # Localize to local timezone and convert to UTC
        local_dt = local_tz.localize(local_naive)
        utc_dt = local_dt.astimezone(pytz.utc)

        # Extract UTC date and interval
        utc_date = utc_dt.date()
        utc_interval = utc_dt.strftime("%H:%M")

        return utc_date, utc_interval

    def _parse_time_intervals(self, hour_interval: str) -> list[str]:
        """Parse time interval string and split into 15-minute intervals."""
        try:
            start_hour, end_hour = hour_interval.split("-")
            start_time = dt.strptime(start_hour.strip(), "%H:%M")
            end_time = dt.strptime(end_hour.strip(), "%H:%M")

            # Split 1-hour intervals into four 15-minute intervals
            if (end_time - start_time).seconds == 3600:
                intervals = [
                    (start_time + timedelta(minutes=15 * i)).strftime("%H:%M") for i in range(4)
                ]
            else:
                intervals = [start_time.strftime("%H:%M")]

            return intervals
        except Exception as e:
            logging.error(f"Error parsing time intervals: {e}")
            return []

    def _create_evaluation_file(self, client_id: int, client_name: str, date_str: str) -> None:
        """Create evaluation file via EvalFileManager."""
        try:
            self.eval_file_manager.create_evaluation_file(client_id, client_name, date_str)
            logging.info(f"Created evaluation file for {client_name} on {date_str}")
        except Exception as e:
            logging.error(f"Error creating evaluation file: {e}")

    def _update_evaluation_files(self, client_name: str, client_id: int, curr_date: Any, tag: str) -> None:
        """Update evaluation files via EvalFileManager."""
        try:
            self.eval_file_manager.update_evaluation_files(client_name, client_id, curr_date, tag)
            logging.info(f"Updated evaluation files for {client_name}")
        except Exception as e:
            logging.error(f"Error updating evaluation files: {e}")

    def _update_evaluation_file(self, client_name: str, client_id: int, plant_name: str, curr_date: Any) -> None:
        """Update single evaluation file via EvalFileManager."""
        try:
            self.eval_file_manager.update_evaluation_files_1plant(client_name, client_id, plant_name, curr_date)
            logging.info(f"Updated evaluation file for {client_name}")
        except Exception as e:
            logging.error(f"Error updating evaluation file: {e}")

    def _update_evaluation_files_1plant(self, client_name: str, client_id: int, plant_name: str, curr_date: Any) -> None:
        """Update evaluation files for single plant via EvalFileManager."""
        try:
            self.eval_file_manager.update_evaluation_files_1plant(client_name, client_id, plant_name, curr_date)
            logging.info(f"Updated plant evaluation for {client_name}")
        except Exception as e:
            logging.error(f"Error updating plant evaluation: {e}")

    def _del_email(self, uid: int) -> None:
        """Delete email by UID using IMAP."""
        try:
            # Convert UID to bytes if needed
            uid_str = str(uid).encode() if isinstance(uid, int) else uid

            # Connect to IMAP server
            mail = imaplib.IMAP4_SSL(IMAP_SERVER)
            mail.login(IMAP_USER, IMAP_PASSWORD)
            mail.select("INBOX")

            # Mark email as deleted
            mail.store(uid_str, "+FLAGS", "\\Deleted")
            mail.expunge()
            mail.close()
            mail.logout()

            logging.info(f"Deleted email with UID {uid}")
        except Exception as e:
            logging.error(f"Error deleting email {uid}: {e}")
